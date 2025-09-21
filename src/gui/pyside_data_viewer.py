from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QPushButton, QTableWidget, QTableWidgetItem,
                               QComboBox, QDateEdit, QTabWidget, QFrame,
                               QHeaderView, QMessageBox, QFileDialog, QScrollArea,
                               QDialog, QTextEdit)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont
from datetime import datetime, timedelta, date
import calendar
import tempfile
import os

class PySideDataViewerWindow(QWidget):
    """Modern PySide6 data viewer for Pomodoro sprints"""

    def __init__(self, parent, db_manager):
        super().__init__()
        self.parent = parent
        self.db_manager = db_manager
        self.current_filter = "day"
        self.current_date = date.today()
        self.chart_images = []  # Track temporary chart image files

        self.init_ui()
        self.apply_styling()
        self.load_data()

    def init_ui(self):
        """Initialize the data viewer UI"""
        self.setWindowTitle("Pomodora - Data Viewer")
        self.setFixedSize(900, 600)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header
        self.create_header(layout)

        # Filter controls
        self.create_filter_controls(layout)

        # Tabs for different views
        self.create_tabs(layout)

        # Action buttons
        self.create_action_buttons(layout)

    def create_header(self, layout):
        """Create header section"""
        header_frame = QFrame()
        header_frame.setObjectName("headerFrame")
        header_layout = QHBoxLayout(header_frame)

        title_label = QLabel("📊 Sprint Data Viewer")
        title_label.setObjectName("titleLabel")
        title_label.setAlignment(Qt.AlignCenter)

        header_layout.addWidget(title_label)
        layout.addWidget(header_frame)

    def create_filter_controls(self, layout):
        """Create filtering controls"""
        filter_frame = QFrame()
        filter_frame.setObjectName("filterFrame")
        filter_layout = QHBoxLayout(filter_frame)

        # View type selector
        filter_layout.addWidget(QLabel("View:"))
        self.view_combo = QComboBox()
        self.view_combo.addItems(["Day", "Week", "Month"])
        self.view_combo.currentTextChanged.connect(self.on_view_changed)
        filter_layout.addWidget(self.view_combo)

        # Date selector
        filter_layout.addWidget(QLabel("Date:"))
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.on_date_changed)
        filter_layout.addWidget(self.date_edit)

        # Navigation buttons
        self.prev_button = QPushButton("◀ Previous")
        self.prev_button.clicked.connect(self.previous_period)
        filter_layout.addWidget(self.prev_button)

        self.next_button = QPushButton("Next ▶")
        self.next_button.clicked.connect(self.next_period)
        filter_layout.addWidget(self.next_button)

        # Today button
        today_button = QPushButton("Today")
        today_button.clicked.connect(self.go_to_today)
        filter_layout.addWidget(today_button)

        filter_layout.addStretch()

        # Summary stats
        self.stats_label = QLabel("Loading...")
        self.stats_label.setObjectName("statsLabel")
        filter_layout.addWidget(self.stats_label)

        layout.addWidget(filter_frame)

    def create_tabs(self, layout):
        """Create tabbed interface"""
        self.tab_widget = QTabWidget()

        # Sprint list tab
        self.create_sprint_list_tab()

        # Summary tab
        self.create_summary_tab()

        layout.addWidget(self.tab_widget)

    def create_sprint_list_tab(self):
        """Create the sprint list table"""
        sprint_widget = QWidget()
        sprint_layout = QVBoxLayout(sprint_widget)

        # Sprint table
        self.sprint_table = QTableWidget()
        self.sprint_table.setColumnCount(7)
        self.sprint_table.setHorizontalHeaderLabels([
            "Date", "Time", "Project", "Category", "Task", "Duration", "Status"
        ])

        # Configure table
        header = self.sprint_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Date
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Time
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Project
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Category
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)           # Task
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Duration
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Status

        self.sprint_table.setAlternatingRowColors(True)
        self.sprint_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Enable sorting but ensure default chronological order (oldest to newest)
        self.sprint_table.setSortingEnabled(True)
        
        # Connect selection change to enable/disable delete button
        self.sprint_table.itemSelectionChanged.connect(self.on_sprint_selection_changed)

        sprint_layout.addWidget(self.sprint_table)
        self.tab_widget.addTab(sprint_widget, "📋 Sprint List")

    def create_summary_tab(self):
        """Create summary statistics tab"""
        summary_widget = QWidget()
        summary_layout = QVBoxLayout(summary_widget)

        # Create scroll area for the summary content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Summary frame (this will be inside the scroll area)
        self.summary_frame = QFrame()
        self.summary_frame.setObjectName("summaryFrame")
        summary_frame_layout = QVBoxLayout(self.summary_frame)

        self.summary_label = QLabel("Loading summary...")
        self.summary_label.setObjectName("summaryText")
        self.summary_label.setAlignment(Qt.AlignTop)
        self.summary_label.setWordWrap(True)
        self.summary_label.setTextFormat(Qt.RichText)  # Enable HTML rendering

        summary_frame_layout.addWidget(self.summary_label)
        
        # Put the summary frame inside the scroll area
        scroll_area.setWidget(self.summary_frame)
        summary_layout.addWidget(scroll_area)

        self.tab_widget.addTab(summary_widget, "📈 Summary")

    def create_action_buttons(self, layout):
        """Create action buttons"""
        button_layout = QHBoxLayout()

        refresh_button = QPushButton("🔄 Refresh")
        refresh_button.clicked.connect(self.load_data)
        button_layout.addWidget(refresh_button)

        export_button = QPushButton("📁 Export to Excel")
        export_button.clicked.connect(self.export_current_view)
        button_layout.addWidget(export_button)

        # Delete sprint button
        self.delete_button = QPushButton("🗑️ Delete Selected Sprint")
        self.delete_button.clicked.connect(self.delete_selected_sprint)
        self.delete_button.setEnabled(False)  # Initially disabled
        button_layout.addWidget(self.delete_button)

        button_layout.addStretch()

        close_button = QPushButton("Close")
        close_button.clicked.connect(self.close)
        button_layout.addWidget(close_button)

        layout.addLayout(button_layout)

    def apply_styling(self):
        """Apply modern styling with theme support"""
        # Detect current theme from parent or settings
        is_dark_mode = self.get_current_theme() == "dark"
        
        if is_dark_mode:
            style = self.get_dark_theme_style()
        else:
            style = self.get_light_theme_style()

        self.setStyleSheet(style)
    
    def get_current_theme(self):
        """Get current theme from parent window or settings"""
        # Try to get theme from parent window first
        if hasattr(self.parent, 'theme_mode'):
            return self.parent.theme_mode
        
        # Fallback to settings
        try:
            from tracking.local_settings import get_local_settings
            settings = get_local_settings()
            return settings.get('theme_mode', 'light')
        except:
            return 'light'
    
    def get_light_theme_style(self):
        """Get light theme stylesheet"""
        return """
        QWidget {
            background: #f8f9fa;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui;
        }

        #headerFrame {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #667eea, stop:1 #764ba2);
            border-radius: 15px;
            padding: 15px;
            margin-bottom: 10px;
        }

        #titleLabel {
            font-size: 20px;
            font-weight: bold;
            color: white;
        }

        #filterFrame {
            background: white;
            border-radius: 10px;
            padding: 15px;
            border: 2px solid #dee2e6;
        }

        #statsLabel {
            font-weight: bold;
            color: #495057;
            background: #e9ecef;
            padding: 8px 12px;
            border-radius: 8px;
        }

        #summaryFrame {
            background: white;
            border-radius: 10px;
            padding: 20px;
            border: 2px solid #dee2e6;
        }

        #summaryText {
            font-size: 14px;
            line-height: 1.6;
            color: #495057;
        }

        QTableWidget {
            background: white;
            border: 2px solid #dee2e6;
            border-radius: 10px;
            gridline-color: #e9ecef;
            font-size: 13px;
        }

        QTableWidget::item {
            padding: 8px;
            border-bottom: 1px solid #e9ecef;
        }

        QTableWidget::item:selected {
            background: #667eea;
            color: white;
        }

        QHeaderView::section {
            background: #f8f9fa;
            border: 1px solid #dee2e6;
            padding: 10px 8px;
            font-weight: bold;
            color: #495057;
        }

        QPushButton {
            background: #667eea;
            color: white;
            border: none;
            border-radius: 8px;
            padding: 8px 16px;
            font-weight: bold;
            font-size: 13px;
        }

        QPushButton:hover {
            background: #5a6fd8;
        }

        QPushButton:pressed {
            background: #4c63d2;
        }

        /* Delete button specific styling */
        QPushButton[text*="Delete"] {
            background: #dc3545;
        }

        QPushButton[text*="Delete"]:hover {
            background: #c82333;
        }

        QPushButton[text*="Delete"]:pressed {
            background: #bd2130;
        }

        QPushButton:disabled {
            background: #6c757d;
            color: #adb5bd;
        }

        QComboBox, QDateEdit {
            background: white;
            border: 2px solid #dee2e6;
            border-radius: 6px;
            padding: 6px 10px;
            font-size: 13px;
        }

        QComboBox:focus, QDateEdit:focus {
            border-color: #667eea;
        }

        QTabWidget::pane {
            border: 2px solid #dee2e6;
            border-radius: 10px;
            background: white;
        }

        QTabBar::tab {
            background: #e9ecef;
            border: none;
            padding: 10px 20px;
            margin-right: 2px;
            border-top-left-radius: 8px;
            border-top-right-radius: 8px;
            font-weight: bold;
        }

        QTabBar::tab:selected {
            background: white;
            color: #667eea;
        }

        QTabBar::tab:hover {
            background: #f8f9fa;
        }
        """
    
    def get_dark_theme_style(self):
        """Get dark theme stylesheet"""
        return """
        QWidget {
            background: #2b2b2b;
            color: #ffffff;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui;
        }

        #headerFrame {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #4a5568, stop:1 #2d3748);
            border-radius: 15px;
            padding: 15px;
            margin-bottom: 10px;
        }

        #titleLabel {
            font-size: 20px;
            font-weight: bold;
            color: white;
        }

        #filterFrame {
            background: #3c3c3c;
            border-radius: 10px;
            padding: 15px;
            border: 2px solid #4a5568;
        }

        #statsLabel {
            font-weight: bold;
            color: #e2e8f0;
            background: #4a5568;
            padding: 8px 12px;
            border-radius: 8px;
        }

        #summaryFrame {
            background: #3c3c3c;
            border-radius: 10px;
            padding: 20px;
            border: 2px solid #4a5568;
        }

        #summaryText {
            font-size: 14px;
            line-height: 1.6;
            color: #e2e8f0;
        }

        QTableWidget {
            background: #3c3c3c;
            border: 2px solid #4a5568;
            border-radius: 10px;
            gridline-color: #4a5568;
            font-size: 13px;
            color: #ffffff;
        }

        QTableWidget::item {
            padding: 8px;
            border-bottom: 1px solid #4a5568;
            background: #3c3c3c;
            color: #ffffff;
        }

        QTableWidget::item:alternate {
            background: #2d3748;
            color: #ffffff;
        }

        QTableWidget::item:selected {
            background: #667eea;
            color: white;
        }

        QHeaderView::section {
            background: #2d3748;
            border: 1px solid #4a5568;
            padding: 10px 8px;
            font-weight: bold;
            color: #e2e8f0;
        }

        QPushButton {
            background: #667eea;
            color: white;
            border: none;
            border-radius: 8px;
            padding: 8px 16px;
            font-weight: bold;
            font-size: 13px;
        }

        QPushButton:hover {
            background: #5a6fd8;
        }

        QPushButton:pressed {
            background: #4c63d2;
        }

        /* Delete button specific styling */
        QPushButton[text*="Delete"] {
            background: #dc3545;
        }

        QPushButton[text*="Delete"]:hover {
            background: #c82333;
        }

        QPushButton[text*="Delete"]:pressed {
            background: #bd2130;
        }

        QPushButton:disabled {
            background: #4a5568;
            color: #718096;
        }

        QComboBox, QDateEdit {
            background: #4a5568;
            color: #ffffff;
            border: 2px solid #667eea;
            border-radius: 6px;
            padding: 6px 10px;
            font-size: 13px;
        }

        QComboBox:focus, QDateEdit:focus {
            border-color: #90cdf4;
        }

        QComboBox::drop-down {
            border: none;
        }

        QComboBox::down-arrow {
            color: #ffffff;
        }

        QTabWidget::pane {
            border: 2px solid #4a5568;
            border-radius: 10px;
            background: #3c3c3c;
        }

        QTabBar::tab {
            background: #2d3748;
            color: #a0aec0;
            border: none;
            padding: 10px 20px;
            margin-right: 2px;
            border-top-left-radius: 8px;
            border-top-right-radius: 8px;
            font-weight: bold;
        }

        QTabBar::tab:selected {
            background: #3c3c3c;
            color: #667eea;
        }

        QTabBar::tab:hover {
            background: #4a5568;
            color: #e2e8f0;
        }

        QLabel {
            color: #e2e8f0;
        }
        """

    def on_view_changed(self, view):
        """Handle view type change"""
        self.current_filter = view.lower()
        self.load_data()

    def on_date_changed(self, qdate):
        """Handle date change"""
        self.current_date = qdate.toPython()
        self.load_data()

    def previous_period(self):
        """Go to previous period"""
        if self.current_filter == "day":
            self.current_date -= timedelta(days=1)
        elif self.current_filter == "week":
            self.current_date -= timedelta(weeks=1)
        elif self.current_filter == "month":
            if self.current_date.month == 1:
                self.current_date = self.current_date.replace(year=self.current_date.year - 1, month=12)
            else:
                self.current_date = self.current_date.replace(month=self.current_date.month - 1)

        self.date_edit.setDate(QDate(self.current_date))
        self.load_data()

    def next_period(self):
        """Go to next period"""
        if self.current_filter == "day":
            self.current_date += timedelta(days=1)
        elif self.current_filter == "week":
            self.current_date += timedelta(weeks=1)
        elif self.current_filter == "month":
            if self.current_date.month == 12:
                self.current_date = self.current_date.replace(year=self.current_date.year + 1, month=1)
            else:
                self.current_date = self.current_date.replace(month=self.current_date.month + 1)

        self.date_edit.setDate(QDate(self.current_date))
        self.load_data()

    def go_to_today(self):
        """Go to today's date"""
        self.current_date = date.today()
        self.date_edit.setDate(QDate.currentDate())
        self.load_data()

    def load_data(self):
        """Load and display sprint data"""
        try:
            sprints = self.get_sprints_for_period()
            self.populate_sprint_table(sprints)
            self.update_summary(sprints)
            self.update_stats_label(sprints)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {str(e)}")

    def get_sprints_for_period(self):
        """Get sprints for the current period"""
        session = self.db_manager.get_session()
        try:
            from tracking.models import Sprint, Project, TaskCategory
            from sqlalchemy.orm import joinedload

            if self.current_filter == "day":
                start_date = datetime.combine(self.current_date, datetime.min.time())
                end_date = start_date + timedelta(days=1)
            elif self.current_filter == "week":
                # Start of week (Monday)
                days_since_monday = self.current_date.weekday()
                start_of_week = self.current_date - timedelta(days=days_since_monday)
                start_date = datetime.combine(start_of_week, datetime.min.time())
                end_date = start_date + timedelta(days=7)
            elif self.current_filter == "month":
                start_date = datetime.combine(
                    self.current_date.replace(day=1), datetime.min.time()
                )
                if self.current_date.month == 12:
                    next_month = start_date.replace(year=start_date.year + 1, month=1)
                else:
                    next_month = start_date.replace(month=start_date.month + 1)
                end_date = next_month

            # Eager load related objects to avoid lazy loading issues
            sprints = session.query(Sprint).options(
                joinedload(Sprint.project),
                joinedload(Sprint.task_category)
            ).filter(
                Sprint.start_time >= start_date,
                Sprint.start_time < end_date
            ).order_by(Sprint.start_time.asc()).all()
            
            # Create detached objects with all data loaded
            detached_sprints = []
            for sprint in sprints:
                # Access all lazy-loaded attributes while session is active
                project_name = sprint.project.name if sprint.project else "Unknown Project"
                task_category_name = sprint.task_category.name if sprint.task_category else "Unknown Category"
                
                # Create a simple data object to avoid session dependency
                sprint_data = type('SprintData', (), {
                    'id': sprint.id,
                    'start_time': sprint.start_time,
                    'end_time': sprint.end_time,
                    'task_description': sprint.task_description,
                    'completed': sprint.completed,
                    'interrupted': sprint.interrupted,
                    'duration_minutes': sprint.duration_minutes,
                    'project_name': project_name,
                    'task_category_name': task_category_name
                })()
                
                detached_sprints.append(sprint_data)
            
            return detached_sprints

        finally:
            session.close()

    def populate_sprint_table(self, sprints):
        """Populate the sprint table with data"""
        # Temporarily disable sorting during population to avoid performance issues
        self.sprint_table.setSortingEnabled(False)
        
        self.sprint_table.setRowCount(len(sprints))
        
        # Store sprints for deletion reference
        self.current_sprints = sprints

        for row, sprint in enumerate(sprints):
            # Date
            date_item = QTableWidgetItem(sprint.start_time.strftime("%Y-%m-%d"))
            # Store sprint index in the first column for easy retrieval
            date_item.setData(Qt.UserRole, row)
            # Store full timestamp for proper sorting
            date_item.setData(Qt.UserRole + 1, sprint.start_time.timestamp())
            self.sprint_table.setItem(row, 0, date_item)

            # Time
            time_item = QTableWidgetItem(sprint.start_time.strftime("%H:%M"))
            # Store full timestamp for proper sorting
            time_item.setData(Qt.UserRole, sprint.start_time.timestamp())
            self.sprint_table.setItem(row, 1, time_item)

            # Project
            project_item = QTableWidgetItem(sprint.project_name)
            self.sprint_table.setItem(row, 2, project_item)

            # Category
            category_item = QTableWidgetItem(sprint.task_category_name)
            self.sprint_table.setItem(row, 3, category_item)

            # Task
            task_item = QTableWidgetItem(sprint.task_description)
            self.sprint_table.setItem(row, 4, task_item)

            # Duration
            if sprint.end_time and sprint.start_time:
                duration = sprint.end_time - sprint.start_time
                duration_mins = int(duration.total_seconds() / 60)
                duration_item = QTableWidgetItem(f"{duration_mins} min")
            else:
                duration_item = QTableWidgetItem("N/A")
            self.sprint_table.setItem(row, 5, duration_item)

            # Status
            status = "✅ Completed" if sprint.completed else ("❌ Interrupted" if sprint.interrupted else "⏸️ Incomplete")
            status_item = QTableWidgetItem(status)
            self.sprint_table.setItem(row, 6, status_item)
        
        # Re-enable sorting and set default chronological order (oldest to newest)
        self.sprint_table.setSortingEnabled(True)
        # Sort by Time column (1) first to get proper time sorting within each date
        # Since we store full timestamp data, this will sort properly by date AND time
        self.sprint_table.sortItems(1, Qt.SortOrder.AscendingOrder)

    def update_summary(self, sprints):
        """Update the summary tab"""
        total_sprints = len(sprints)
        completed_sprints = len([s for s in sprints if s.completed])
        interrupted_sprints = len([s for s in sprints if s.interrupted])

        total_time = 0
        projects = {}
        categories = {}
        task_descriptions = {}

        for sprint in sprints:
            if sprint.end_time and sprint.start_time:
                duration = sprint.end_time - sprint.start_time
                total_time += duration.total_seconds() / 60

            project = sprint.project_name
            if project not in projects:
                projects[project] = 0
            projects[project] += 1
            
            category = sprint.task_category_name
            if category not in categories:
                categories[category] = 0
            categories[category] += 1
            
            # Track task descriptions
            task_desc = sprint.task_description.strip() if sprint.task_description else "No Description"
            if task_desc not in task_descriptions:
                task_descriptions[task_desc] = 0
            task_descriptions[task_desc] += 1

        # Calculate completion rate
        completion_rate = (completed_sprints / total_sprints * 100) if total_sprints > 0 else 0

        # Format time
        hours = int(total_time / 60)
        minutes = int(total_time % 60)

        # Generate summary text
        period_name = self.current_filter.title()
        period_date = self.current_date.strftime("%Y-%m-%d")

        summary_text = f"""
<h2>📊 {period_name} Summary - {period_date}</h2>

<h3>🎯 Sprint Statistics</h3>
<ul>
<li><b>Total Sprints:</b> {total_sprints}</li>
<li><b>Completed:</b> {completed_sprints} ({completion_rate:.1f}%)</li>
<li><b>Interrupted:</b> {interrupted_sprints}</li>
<li><b>Total Focus Time:</b> {hours}h {minutes}m</li>
</ul>

<h3>📋 Projects Breakdown</h3>
<ul>
"""

        if projects:
            for project, count in sorted(projects.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_sprints * 100) if total_sprints > 0 else 0
                summary_text += f"<li><b>{project}:</b> {count} sprints ({percentage:.1f}%)</li>\n"
        else:
            summary_text += "<li><i>No projects found</i></li>\n"

        summary_text += "</ul>"

        # Add project pie chart right after project breakdown
        if projects and len(projects) > 1:
            # Clean up any previous chart images
            self.cleanup_chart_images()
            project_chart_path = self.create_pie_chart(projects, "Projects Distribution", total_sprints)
            if project_chart_path:
                summary_text += f"""
<p style="text-align: center; margin: 20px 0;">
<img src="file://{project_chart_path}" alt="Projects Pie Chart" style="max-width: 450px; height: auto; border-radius: 8px;">
</p>
"""

        summary_text += f"""

<h3>🏷️ Task Categories Breakdown</h3>
<ul>
"""

        if categories:
            for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_sprints * 100) if total_sprints > 0 else 0
                summary_text += f"<li><b>{category}:</b> {count} sprints ({percentage:.1f}%)</li>\n"
        else:
            summary_text += "<li><i>No task categories found</i></li>\n"

        summary_text += "</ul>"

        # Add category pie chart right after category breakdown
        if categories and len(categories) > 1:
            category_chart_path = self.create_pie_chart(categories, "Task Categories Distribution", total_sprints)
            if category_chart_path:
                summary_text += f"""
<p style="text-align: center; margin: 20px 0;">
<img src="file://{category_chart_path}" alt="Categories Pie Chart" style="max-width: 450px; height: auto; border-radius: 8px;">
</p>
"""

        # Task Description Analysis - only show if there are frequent task descriptions (>10%)
        frequent_tasks = {}
        other_tasks_count = 0
        
        if task_descriptions and total_sprints > 0:
            for task_desc, count in task_descriptions.items():
                percentage = (count / total_sprints) * 100
                if percentage > 10.0:
                    frequent_tasks[task_desc] = count
                else:
                    other_tasks_count += count
        
        # Only show task description breakdown if there's at least one frequent task
        if frequent_tasks:
            summary_text += f"""

<h3>📝 Task Descriptions Breakdown</h3>
<ul>
"""
            
            # Sort frequent tasks by count (descending)
            for task_desc, count in sorted(frequent_tasks.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_sprints * 100) if total_sprints > 0 else 0
                # Truncate long task descriptions for display
                display_desc = task_desc if len(task_desc) <= 50 else f"{task_desc[:47]}..."
                summary_text += f"<li><b>{display_desc}:</b> {count} sprints ({percentage:.1f}%)</li>\n"
            
            # Add "Other" category if there are remaining tasks
            if other_tasks_count > 0:
                other_percentage = (other_tasks_count / total_sprints * 100) if total_sprints > 0 else 0
                summary_text += f"<li><b>Other:</b> {other_tasks_count} sprints ({other_percentage:.1f}%)</li>\n"
            
            summary_text += "</ul>"
            
            # Create task description pie chart with "Other" consolidation
            chart_data = frequent_tasks.copy()
            if other_tasks_count > 0:
                chart_data["Other"] = other_tasks_count
                
            if len(chart_data) > 1:
                task_chart_path = self.create_pie_chart(chart_data, "Task Descriptions Distribution", total_sprints)
                if task_chart_path:
                    summary_text += f"""
<p style="text-align: center; margin: 20px 0;">
<img src="file://{task_chart_path}" alt="Task Descriptions Pie Chart" style="max-width: 450px; height: auto; border-radius: 8px;">
</p>
"""

        # Add time-based line graphs based on view type
        if total_sprints > 0:
            if self.current_filter == "month":
                # Add weekly sprint count graph for monthly view
                weekly_chart_path = self.create_weekly_line_chart(sprints)
                if weekly_chart_path:
                    summary_text += f"""
<h3>📈 Sprint Counts by Week</h3>
<p style="text-align: center; margin: 20px 0;">
<img src="file://{weekly_chart_path}" alt="Weekly Sprint Counts" style="max-width: 600px; height: auto; border-radius: 8px;">
</p>
"""
            elif self.current_filter == "week":
                # Add daily sprint count graph for weekly view
                daily_chart_path = self.create_daily_line_chart(sprints)
                if daily_chart_path:
                    summary_text += f"""
<h3>📈 Sprint Counts by Day</h3>
<p style="text-align: center; margin: 20px 0;">
<img src="file://{daily_chart_path}" alt="Daily Sprint Counts" style="max-width: 600px; height: auto; border-radius: 8px;">
</p>
"""

        if total_sprints == 0:
            summary_text += "\n<p><i>No sprints found for this period.</i></p>"

        self.summary_label.setText(summary_text)


    def create_pie_chart(self, data_dict, title, total):
        """Create a graphical pie chart using matplotlib with theme support"""
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend
            
            if not data_dict or total == 0:
                return None
                
            # Detect current theme
            is_dark_theme = self.get_current_theme() == "dark"
            
            # Sort by count (descending)
            sorted_data = sorted(data_dict.items(), key=lambda x: x[1], reverse=True)
            
            # Prepare data
            labels = []
            sizes = []
            colors = []
            
            # Define theme-appropriate color palettes
            if is_dark_theme:
                # Brighter, more saturated colors for dark theme
                color_palette = [
                    '#FF7B7B', '#5EDCE4', '#55C7E1', '#A6DEB4', '#FFFA87', 
                    '#EDA0DD', '#A8E8D8', '#F7EC6F', '#CB9FCE', '#95D1F9'
                ]
                bg_color = '#2b2b2b'
                text_color = '#ffffff'
                title_color = '#ffffff'
            else:
                # Standard colors for light theme
                color_palette = [
                    '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
                    '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9'
                ]
                bg_color = '#ffffff'
                text_color = '#000000'
                title_color = '#000000'
            
            for i, (name, count) in enumerate(sorted_data):
                percentage = (count / total) * 100
                labels.append(f'{name}\n({count} sprints, {percentage:.1f}%)')
                sizes.append(count)
                colors.append(color_palette[i % len(color_palette)])
            
            # Create figure and pie chart with larger size
            fig, ax = plt.subplots(figsize=(10, 10))
            fig.patch.set_facecolor(bg_color)
            
            # Create pie chart
            wedges, texts, autotexts = ax.pie(
                sizes, 
                labels=labels,
                colors=colors,
                autopct='%1.1f%%',
                startangle=90,
                explode=[0.08] * len(sizes),  # Slightly more separation for clarity
                textprops={'fontsize': 12, 'fontweight': 'bold', 'color': text_color}  # Larger text
            )
            
            # Customize appearance with larger fonts
            ax.set_title(title, fontsize=20, fontweight='bold', pad=30, color=title_color)
            
            # Make percentage text more readable and larger
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(14)  # Larger percentage text
                autotext.set_bbox(dict(boxstyle="round,pad=0.3", facecolor='black', alpha=0.7))
            
            # Make label text larger and more readable
            for text in texts:
                text.set_fontsize(12)  # Larger label text
                text.set_fontweight('bold')
                text.set_color(text_color)
            
            # Equal aspect ratio ensures that pie is drawn as a circle
            ax.axis('equal')
            ax.set_facecolor(bg_color)
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Save the chart with theme-appropriate background
            plt.tight_layout()
            plt.savefig(temp_path, dpi=200, bbox_inches='tight', 
                       facecolor=bg_color, edgecolor='none')
            plt.close(fig)
            
            # Track the file for cleanup
            self.chart_images.append(temp_path)
            
            return temp_path
            
        except ImportError:
            return None
        except Exception as e:
            print(f"Error creating pie chart: {e}")
            return None

    def create_weekly_line_chart(self, sprints):
        """Create a line chart showing sprint counts by week for monthly view"""
        try:
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend

            if not sprints:
                return None

            # Detect current theme
            is_dark_theme = self.get_current_theme() == "dark"

            # Group sprints by week
            from collections import defaultdict
            import calendar

            weekly_counts = defaultdict(int)

            # Get start of month for reference
            month_start = self.current_date.replace(day=1)

            for sprint in sprints:
                sprint_date = sprint.start_time.date()
                # Calculate which week of the month this sprint belongs to
                week_start = sprint_date - timedelta(days=sprint_date.weekday())
                weekly_counts[week_start] += 1

            if not weekly_counts:
                return None

            # Sort weeks and prepare data
            sorted_weeks = sorted(weekly_counts.keys())
            week_labels = []
            counts = []

            for week_start in sorted_weeks:
                # Format week label as "Week of Mon DD"
                week_labels.append(week_start.strftime("Week of %b %d"))
                counts.append(weekly_counts[week_start])

            # Set theme colors
            if is_dark_theme:
                bg_color = '#2b2b2b'
                text_color = '#ffffff'
                line_color = '#5EDCE4'
                grid_color = '#4a5568'
                title_color = '#ffffff'
            else:
                bg_color = '#ffffff'
                text_color = '#000000'
                line_color = '#4ECDC4'
                grid_color = '#dee2e6'
                title_color = '#000000'

            # Create figure
            fig, ax = plt.subplots(figsize=(12, 6))
            fig.patch.set_facecolor(bg_color)
            ax.set_facecolor(bg_color)

            # Create line plot
            ax.plot(range(len(week_labels)), counts,
                   color=line_color, linewidth=3, marker='o',
                   markersize=8, markerfacecolor=line_color,
                   markeredgecolor='white', markeredgewidth=2)

            # Customize appearance
            ax.set_xlabel('Week', fontsize=14, fontweight='bold', color=text_color)
            ax.set_ylabel('Number of Sprints', fontsize=14, fontweight='bold', color=text_color)
            ax.set_title(f'Sprint Counts by Week - {self.current_date.strftime("%B %Y")}',
                        fontsize=16, fontweight='bold', pad=20, color=title_color)

            # Set x-axis labels
            ax.set_xticks(range(len(week_labels)))
            ax.set_xticklabels(week_labels, rotation=45, ha='right',
                              fontsize=12, color=text_color)

            # Style y-axis
            ax.tick_params(axis='y', labelsize=12, colors=text_color)

            # Add grid
            ax.grid(True, alpha=0.3, color=grid_color)

            # Ensure y-axis starts at 0 and uses integer ticks
            ax.set_ylim(bottom=0)
            max_count = max(counts) if counts else 1
            ax.set_yticks(range(0, max_count + 2))

            # Add value labels on data points
            for i, count in enumerate(counts):
                ax.annotate(str(count), (i, count),
                           textcoords="offset points", xytext=(0,10),
                           ha='center', fontsize=12, fontweight='bold',
                           color=text_color,
                           bbox=dict(boxstyle="round,pad=0.3",
                                   facecolor=line_color, alpha=0.7))

            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_path = temp_file.name
            temp_file.close()

            plt.tight_layout()
            plt.savefig(temp_path, dpi=150, bbox_inches='tight',
                       facecolor=bg_color, edgecolor='none')
            plt.close(fig)

            # Track the file for cleanup
            self.chart_images.append(temp_path)

            return temp_path

        except ImportError:
            return None
        except Exception as e:
            print(f"Error creating weekly line chart: {e}")
            return None

    def create_daily_line_chart(self, sprints):
        """Create a line chart showing sprint counts by day for weekly view"""
        try:
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend

            if not sprints:
                return None

            # Detect current theme
            is_dark_theme = self.get_current_theme() == "dark"

            # Group sprints by day
            from collections import defaultdict

            daily_counts = defaultdict(int)

            # Get week boundaries
            days_since_monday = self.current_date.weekday()
            week_start = self.current_date - timedelta(days=days_since_monday)

            for sprint in sprints:
                sprint_date = sprint.start_time.date()
                daily_counts[sprint_date] += 1

            # Create weekday data (Monday through Friday only)
            week_days = []
            counts = []
            day_names = []

            for i in range(5):  # Monday (0) through Friday (4)
                day = week_start + timedelta(days=i)
                week_days.append(day)
                counts.append(daily_counts.get(day, 0))
                day_names.append(day.strftime("%a %m/%d"))  # e.g., "Mon 12/25"

            # Set theme colors
            if is_dark_theme:
                bg_color = '#2b2b2b'
                text_color = '#ffffff'
                line_color = '#FF7B7B'
                grid_color = '#4a5568'
                title_color = '#ffffff'
            else:
                bg_color = '#ffffff'
                text_color = '#000000'
                line_color = '#FF6B6B'
                grid_color = '#dee2e6'
                title_color = '#000000'

            # Create figure
            fig, ax = plt.subplots(figsize=(12, 6))
            fig.patch.set_facecolor(bg_color)
            ax.set_facecolor(bg_color)

            # Create line plot
            ax.plot(range(len(day_names)), counts,
                   color=line_color, linewidth=3, marker='o',
                   markersize=8, markerfacecolor=line_color,
                   markeredgecolor='white', markeredgewidth=2)

            # Customize appearance
            ax.set_xlabel('Day of Week', fontsize=14, fontweight='bold', color=text_color)
            ax.set_ylabel('Number of Sprints', fontsize=14, fontweight='bold', color=text_color)
            ax.set_title(f'Sprint Counts by Day - Week of {week_start.strftime("%B %d, %Y")}',
                        fontsize=16, fontweight='bold', pad=20, color=title_color)

            # Set x-axis labels
            ax.set_xticks(range(len(day_names)))
            ax.set_xticklabels(day_names, fontsize=12, color=text_color)

            # Style y-axis
            ax.tick_params(axis='y', labelsize=12, colors=text_color)

            # Add grid
            ax.grid(True, alpha=0.3, color=grid_color)

            # Ensure y-axis starts at 0 and uses integer ticks
            ax.set_ylim(bottom=0)
            max_count = max(counts) if any(counts) else 1
            ax.set_yticks(range(0, max_count + 2))

            # Add value labels on data points
            for i, count in enumerate(counts):
                if count > 0:  # Only show labels for non-zero values
                    ax.annotate(str(count), (i, count),
                               textcoords="offset points", xytext=(0,10),
                               ha='center', fontsize=12, fontweight='bold',
                               color=text_color,
                               bbox=dict(boxstyle="round,pad=0.3",
                                       facecolor=line_color, alpha=0.7))

            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_path = temp_file.name
            temp_file.close()

            plt.tight_layout()
            plt.savefig(temp_path, dpi=150, bbox_inches='tight',
                       facecolor=bg_color, edgecolor='none')
            plt.close(fig)

            # Track the file for cleanup
            self.chart_images.append(temp_path)

            return temp_path

        except ImportError:
            return None
        except Exception as e:
            print(f"Error creating daily line chart: {e}")
            return None

    def cleanup_chart_images(self):
        """Clean up temporary chart image files"""
        for image_path in self.chart_images:
            try:
                if os.path.exists(image_path):
                    os.unlink(image_path)
            except Exception:
                pass
        self.chart_images = []

    def closeEvent(self, event):
        """Clean up when window is closed"""
        self.cleanup_chart_images()
        super().closeEvent(event)

    def update_stats_label(self, sprints):
        """Update the stats label in the header"""
        total = len(sprints)
        completed = len([s for s in sprints if s.completed])

        if total > 0:
            total_time = sum(
                (s.end_time - s.start_time).total_seconds() / 60
                for s in sprints
                if s.end_time and s.start_time
            )
            hours = int(total_time / 60)
            minutes = int(total_time % 60)

            self.stats_label.setText(
                f"📊 {total} sprints • {completed} completed • {hours}h {minutes}m total"
            )
        else:
            self.stats_label.setText("📊 No data for this period")

    def export_current_view(self):
        """Export current view to Excel"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Current View",
                f"pomodora_{self.current_filter}_{self.current_date.strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                sprints = self.get_sprints_for_period()
                self.export_sprints_to_excel(sprints, file_path)
                QMessageBox.information(self, "Export Complete",
                                      f"Data exported successfully to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data:\n{str(e)}")

    def export_sprints_to_excel(self, sprints, file_path):
        """Export sprints to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Sprints_{self.current_filter.title()}"

            # Headers
            headers = ["Date", "Time", "Project", "Category", "Task", "Duration (min)", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row, sprint in enumerate(sprints, 2):
                ws.cell(row=row, column=1, value=sprint.start_time.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=sprint.start_time.strftime("%H:%M"))
                ws.cell(row=row, column=3, value=sprint.project_name)
                ws.cell(row=row, column=4, value=sprint.task_category_name)
                ws.cell(row=row, column=5, value=sprint.task_description)

                if sprint.end_time and sprint.start_time:
                    duration = (sprint.end_time - sprint.start_time).total_seconds() / 60
                    ws.cell(row=row, column=6, value=f"{int(duration)}")
                else:
                    ws.cell(row=row, column=6, value="N/A")

                status = "Completed" if sprint.completed else ("Interrupted" if sprint.interrupted else "Incomplete")
                ws.cell(row=row, column=7, value=status)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

        except ImportError:
            raise Exception("openpyxl library not found. Please install it with: pip install openpyxl")

    def on_sprint_selection_changed(self):
        """Handle sprint table selection changes"""
        selected_rows = self.sprint_table.selectionModel().selectedRows()
        self.delete_button.setEnabled(len(selected_rows) > 0)

    def delete_selected_sprint(self):
        """Delete the currently selected sprint"""
        selected_rows = self.sprint_table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select a sprint to delete.")
            return
            
        # Get the selected sprint
        row = selected_rows[0].row()
        if row >= len(self.current_sprints):
            QMessageBox.warning(self, "Error", "Invalid sprint selection.")
            return
            
        sprint = self.current_sprints[row]
        
        # Confirm deletion with custom dialog
        confirmation_dialog = QDialog(self)
        confirmation_dialog.setWindowTitle("Confirm Deletion")
        confirmation_dialog.setFixedSize(450, 250)
        confirmation_dialog.setModal(True)
        
        # Apply theme-aware styling to dialog
        is_dark_theme = self.get_current_theme() == 'dark'
        
        if is_dark_theme:
            dialog_style = """
            QDialog {
                background: #1a202c;
                color: #e2e8f0;
            }
            """
            details_bg = "#2d3748"
            details_color = "#e2e8f0"
            warning_color = "#718096"
            cancel_style = "padding: 8px 20px; font-size: 12px; background: #4a5568; color: #e2e8f0; border: none; border-radius: 4px;"
        else:
            dialog_style = """
            QDialog {
                background: #ffffff;
                color: #000000;
            }
            """
            details_bg = "#f8f9fa"
            details_color = "#212529"
            warning_color = "#6c757d"
            cancel_style = "padding: 8px 20px; font-size: 12px; background: #6c757d; color: white; border: none; border-radius: 4px;"
        
        confirmation_dialog.setStyleSheet(dialog_style)
        
        layout = QVBoxLayout(confirmation_dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Warning message
        warning_label = QLabel("⚠️ Are you sure you want to delete this sprint?")
        warning_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #dc3545;")
        layout.addWidget(warning_label)
        
        # Sprint details
        details_text = (
            f"Project: {sprint.project_name}\n"
            f"Task: {sprint.task_description}\n"
            f"Date: {sprint.start_time.strftime('%Y-%m-%d %H:%M')}"
        )
        details_label = QLabel(details_text)
        details_label.setStyleSheet(f"font-size: 12px; padding: 10px; background: {details_bg}; color: {details_color}; border-radius: 5px;")
        details_label.setWordWrap(True)
        layout.addWidget(details_label)
        
        # Final warning
        final_warning = QLabel("This action cannot be undone.")
        final_warning.setStyleSheet(f"font-size: 12px; font-style: italic; color: {warning_color};")
        layout.addWidget(final_warning)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setStyleSheet(cancel_style)
        cancel_button.clicked.connect(confirmation_dialog.reject)
        button_layout.addWidget(cancel_button)
        
        delete_button = QPushButton("Delete Sprint")
        delete_button.setStyleSheet("padding: 8px 20px; font-size: 12px; background: #dc3545; color: white; border: none; border-radius: 4px;")
        delete_button.clicked.connect(confirmation_dialog.accept)
        button_layout.addWidget(delete_button)
        
        layout.addLayout(button_layout)
        
        reply = confirmation_dialog.exec()
        
        if reply == QDialog.DialogCode.Accepted:
            try:
                # Delete the sprint from database
                success = self.db_manager.delete_sprint(sprint.id)
                
                if success:
                    # Create a properly sized success dialog
                    success_dialog = QDialog(self)
                    success_dialog.setWindowTitle("Success")
                    success_dialog.setFixedSize(350, 150)
                    success_dialog.setModal(True)
                    
                    # Apply theme-aware styling
                    is_dark_theme = self.get_current_theme() == 'dark'
                    
                    if is_dark_theme:
                        dialog_style = """
                        QDialog {
                            background: #1a202c;
                            color: #e2e8f0;
                        }
                        """
                        ok_button_style = "padding: 8px 20px; font-size: 12px; min-width: 60px; background: #667eea; color: white; border: none; border-radius: 4px;"
                    else:
                        dialog_style = """
                        QDialog {
                            background: #ffffff;
                            color: #000000;
                        }
                        """
                        ok_button_style = "padding: 8px 20px; font-size: 12px; min-width: 60px; background: #667eea; color: white; border: none; border-radius: 4px;"
                    
                    success_dialog.setStyleSheet(dialog_style)
                    
                    layout = QVBoxLayout(success_dialog)
                    layout.setSpacing(15)
                    layout.setContentsMargins(20, 20, 20, 20)
                    
                    # Success message
                    success_label = QLabel("✅ Sprint deleted successfully!")
                    success_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #28a745;")
                    success_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    layout.addWidget(success_label)
                    
                    # OK button
                    button_layout = QHBoxLayout()
                    button_layout.addStretch()
                    
                    ok_button = QPushButton("OK")
                    ok_button.setStyleSheet(ok_button_style)
                    ok_button.clicked.connect(success_dialog.accept)
                    button_layout.addWidget(ok_button)
                    button_layout.addStretch()
                    
                    layout.addLayout(button_layout)
                    
                    success_dialog.exec()
                    # Refresh the view
                    self.load_data()
                    
                    # Notify parent window to update stats if deletion affects today's sprints
                    if hasattr(self.parent, 'update_stats'):
                        self.parent.update_stats()
                else:
                    QMessageBox.warning(self, "Error", "Failed to delete sprint.")
                    
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while deleting the sprint:\n{str(e)}")