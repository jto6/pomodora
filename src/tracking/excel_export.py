import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_GENERAL
from datetime import datetime, timedelta
import calendar
from sqlalchemy import and_
from .database_manager_unified import UnifiedDatabaseManager
from .models import Sprint, Project
from utils.progress_wrapper import with_progress, ProgressCapableMixin

class ExcelExporter(ProgressCapableMixin):
    def __init__(self, db_manager: UnifiedDatabaseManager):
        self.db_manager = db_manager

        # Define colors
        self.colors = {
            'header': 'D9E2F3',
            'project_header': 'E2EFDA',
            'total': 'FFF2CC',
            'border': '000000'
        }

    @with_progress("Exporting Monthly Data", "Creating Excel workbook with monthly analysis...")
    def export_month(self, year: int, month: int, filename: str):
        """Export data for a specific month in template format"""
        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Get month name
        month_name = calendar.month_name[month]

        # Create month summary sheet
        self.create_month_summary_sheet(wb, year, month)

        # Create weekly sheets
        self.create_weekly_sheets(wb, year, month)

        # Save workbook
        wb.save(filename)

    def create_month_summary_sheet(self, wb: openpyxl.Workbook, year: int, month: int):
        """Create month summary sheet"""
        month_name = calendar.month_name[month]
        ws = wb.create_sheet(f"{month_name} {year}")

        # Get data for the month
        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)

        session = self.db_manager.get_session()
        try:
            # Get all sprints for the month
            sprints = session.query(Sprint).filter(
                and_(
                    Sprint.start_time >= start_date,
                    Sprint.start_time <= end_date
                )
            ).order_by(Sprint.start_time).all()

            # Get all projects
            projects = session.query(Project).all()
            project_colors = {p.name: p.color for p in projects}

            # Set up headers
            ws['A1'] = f"Pomodoro Activity Tracker - {month_name} {year}"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:G1')

            # Create project summary
            row = 3
            ws[f'A{row}'] = "Project Summary"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            # Headers
            headers = ['Project', 'Total Sprints', 'Completed', 'Total Minutes', 'Avg Duration']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            row += 1

            # Project data
            project_stats = self.calculate_project_stats(sprints)
            for project_name, stats in project_stats.items():
                ws.cell(row=row, column=1, value=project_name)
                ws.cell(row=row, column=2, value=stats['total'])
                ws.cell(row=row, column=3, value=stats['completed'])
                ws.cell(row=row, column=4, value=stats['minutes'])
                ws.cell(row=row, column=5, value=f"{stats['avg']:.1f}")

                # Color code project name if color is available
                if project_name in project_colors:
                    try:
                        color = project_colors[project_name].replace('#', '')
                        ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    except:
                        pass

                row += 1

            # Daily breakdown
            row += 2
            ws[f'A{row}'] = "Daily Activity"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            # Daily headers
            daily_headers = ['Date', 'Day', 'Total Sprints', 'Completed', 'Total Minutes']
            for col, header in enumerate(daily_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            row += 1

            # Daily data
            daily_stats = self.calculate_daily_stats(sprints, year, month)
            for date, stats in daily_stats.items():
                ws.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=date.strftime('%A'))
                ws.cell(row=row, column=3, value=stats['total'])
                ws.cell(row=row, column=4, value=stats['completed'])
                ws.cell(row=row, column=5, value=stats['minutes'])
                row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        finally:
            session.close()

    def create_weekly_sheets(self, wb: openpyxl.Workbook, year: int, month: int):
        """Create weekly breakdown sheets"""
        # Get all weeks in the month
        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)

        # Find all Mondays in the month
        current_date = start_date
        week_num = 1

        while current_date.month == month:
            # Find Monday of this week
            days_since_monday = current_date.weekday()
            monday = current_date - timedelta(days=days_since_monday)

            # Find Sunday of this week
            sunday = monday + timedelta(days=6)

            # Create week sheet
            self.create_week_sheet(wb, monday, sunday, f"Week {week_num}")

            # Move to next week
            current_date = sunday + timedelta(days=1)
            week_num += 1

            # Safety check
            if week_num > 6:  # No month has more than 6 weeks
                break

    def create_week_sheet(self, wb: openpyxl.Workbook, start_date: datetime, end_date: datetime, sheet_name: str):
        """Create a weekly activity sheet"""
        ws = wb.create_sheet(sheet_name)

        session = self.db_manager.get_session()
        try:
            # Get sprints for this week
            sprints = session.query(Sprint).filter(
                and_(
                    Sprint.start_time >= start_date,
                    Sprint.start_time <= end_date + timedelta(days=1)
                )
            ).order_by(Sprint.start_time).all()

            # Sheet header
            ws['A1'] = f"Week of {start_date.strftime('%B %d, %Y')}"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')

            # Table headers
            row = 3
            headers = ['Date', 'Time', 'Project', 'Task Description', 'Duration (min)', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            row += 1

            # Sprint data
            for sprint in sprints:
                ws.cell(row=row, column=1, value=sprint.start_time.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=sprint.start_time.strftime('%H:%M'))
                ws.cell(row=row, column=3, value=sprint.project_name)
                ws.cell(row=row, column=4, value=sprint.task_description)
                ws.cell(row=row, column=5, value=sprint.duration_minutes or 0)

                status = "Completed" if sprint.completed else ("Interrupted" if sprint.interrupted else "In Progress")
                ws.cell(row=row, column=6, value=status)

                row += 1

            # Week summary
            if sprints:
                row += 1
                total_sprints = len(sprints)
                completed_sprints = sum(1 for s in sprints if s.completed)
                total_minutes = sum(s.duration_minutes or 0 for s in sprints)

                ws.cell(row=row, column=3, value="TOTALS:").font = Font(bold=True)
                ws.cell(row=row, column=4, value=f"{total_sprints} sprints, {completed_sprints} completed")
                ws.cell(row=row, column=5, value=total_minutes).font = Font(bold=True)

                # Highlight totals row
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                # Skip merged cells
                if hasattr(column[0], 'column_letter'):
                    column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

        finally:
            session.close()

    @with_progress("Exporting Date Range", "Generating Excel report for selected date range...")
    def export_date_range(self, start_date: datetime, end_date: datetime, filename: str):
        """Export data for a specific date range"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sprint Data"

        session = self.db_manager.get_session()
        try:
            # Get sprints in date range
            sprints = session.query(Sprint).filter(
                and_(
                    Sprint.start_time >= start_date,
                    Sprint.start_time < end_date
                )
            ).order_by(Sprint.start_time).all()

            # Headers
            headers = ['Date', 'Time', 'Project', 'Task Description', 'Duration (min)', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')

            # Data
            for row, sprint in enumerate(sprints, 2):
                ws.cell(row=row, column=1, value=sprint.start_time.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=sprint.start_time.strftime('%H:%M'))
                ws.cell(row=row, column=3, value=sprint.project_name)
                ws.cell(row=row, column=4, value=sprint.task_description)
                ws.cell(row=row, column=5, value=sprint.duration_minutes or 0)

                status = "Completed" if sprint.completed else ("Interrupted" if sprint.interrupted else "In Progress")
                ws.cell(row=row, column=6, value=status)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                # Skip merged cells
                if hasattr(column[0], 'column_letter'):
                    column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

        finally:
            session.close()

        wb.save(filename)

    def calculate_project_stats(self, sprints):
        """Calculate statistics by project"""
        stats = {}

        for sprint in sprints:
            project = sprint.project_name
            if project not in stats:
                stats[project] = {'total': 0, 'completed': 0, 'minutes': 0}

            stats[project]['total'] += 1
            if sprint.completed:
                stats[project]['completed'] += 1
            stats[project]['minutes'] += sprint.duration_minutes or 0

        # Calculate averages
        for project, data in stats.items():
            data['avg'] = data['minutes'] / data['total'] if data['total'] > 0 else 0

        return stats

    def calculate_daily_stats(self, sprints, year: int, month: int):
        """Calculate statistics by day"""
        stats = {}

        # Initialize all days in month
        last_day = calendar.monthrange(year, month)[1]
        for day in range(1, last_day + 1):
            date = datetime(year, month, day)
            stats[date] = {'total': 0, 'completed': 0, 'minutes': 0}

        # Populate with sprint data
        for sprint in sprints:
            date = sprint.start_time.replace(hour=0, minute=0, second=0, microsecond=0)
            if date in stats:
                stats[date]['total'] += 1
                if sprint.completed:
                    stats[date]['completed'] += 1
                stats[date]['minutes'] += sprint.duration_minutes or 0

        return stats

    @with_progress("Exporting All Data", "Creating comprehensive Excel workbook with all sprint data...")
    def export_all_data(self, filename: str):
        """Export all data in a comprehensive workbook"""
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        session = self.db_manager.get_session()
        try:
            # Get all sprints
            all_sprints = session.query(Sprint).order_by(Sprint.start_time).all()

            if not all_sprints:
                # Create empty sheet if no data
                ws = wb.create_sheet("No Data")
                ws['A1'] = "No sprint data found"
                wb.save(filename)
                return

            # Create overview sheet
            self.create_overview_sheet(wb, all_sprints)

            # Create all sprints sheet
            self.create_all_sprints_sheet(wb, all_sprints)

            # Create project summary sheet
            self.create_project_summary_sheet(wb, all_sprints)

            # Get date range for monthly sheets
            first_sprint = min(all_sprints, key=lambda s: s.start_time)
            last_sprint = max(all_sprints, key=lambda s: s.start_time)

            # Create monthly sheets for each month with data
            current_date = first_sprint.start_time.replace(day=1)
            end_date = last_sprint.start_time.replace(day=1)

            while current_date <= end_date:
                # Check if this month has data
                month_sprints = [s for s in all_sprints
                               if s.start_time.year == current_date.year and
                                  s.start_time.month == current_date.month]

                if month_sprints:
                    self.create_month_summary_sheet(wb, current_date.year, current_date.month)

                # Move to next month
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

        finally:
            session.close()

        wb.save(filename)

    def create_overview_sheet(self, wb: openpyxl.Workbook, all_sprints):
        """Create overview summary sheet"""
        ws = wb.create_sheet("Overview")

        # Title
        ws['A1'] = "Pomodoro Activity Overview"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        row = 3

        # Total statistics
        total_sprints = len(all_sprints)
        completed_sprints = sum(1 for s in all_sprints if s.completed)
        total_minutes = sum(s.duration_minutes or 0 for s in all_sprints)

        stats = [
            ("Total Sprints", total_sprints),
            ("Completed Sprints", completed_sprints),
            ("Total Time (minutes)", total_minutes),
            ("Total Time (hours)", f"{total_minutes / 60:.1f}"),
            ("Completion Rate", f"{(completed_sprints / total_sprints * 100):.1f}%" if total_sprints > 0 else "0%")
        ]

        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Date range
        row += 1
        if all_sprints:
            first_date = min(s.start_time for s in all_sprints).strftime('%Y-%m-%d')
            last_date = max(s.start_time for s in all_sprints).strftime('%Y-%m-%d')
            ws.cell(row=row, column=1, value="Date Range").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{first_date} to {last_date}")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_all_sprints_sheet(self, wb: openpyxl.Workbook, all_sprints):
        """Create sheet with all sprint data"""
        ws = wb.create_sheet("All Sprints")

        # Headers
        headers = ['Date', 'Time', 'Project', 'Task Description', 'Duration (min)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')

        # Data
        for row, sprint in enumerate(all_sprints, 2):
            ws.cell(row=row, column=1, value=sprint.start_time.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=sprint.start_time.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=sprint.project_name)
            ws.cell(row=row, column=4, value=sprint.task_description)
            ws.cell(row=row, column=5, value=sprint.duration_minutes or 0)

            status = "Completed" if sprint.completed else ("Interrupted" if sprint.interrupted else "In Progress")
            ws.cell(row=row, column=6, value=status)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_project_summary_sheet(self, wb: openpyxl.Workbook, all_sprints):
        """Create project summary sheet"""
        ws = wb.create_sheet("Project Summary")

        # Title
        ws['A1'] = "Project Summary"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        row = 3
        headers = ['Project', 'Total Sprints', 'Completed', 'Total Minutes', 'Avg Duration', 'Completion Rate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        row += 1

        # Project data
        project_stats = self.calculate_project_stats(all_sprints)
        for project_name, stats in project_stats.items():
            completion_rate = (stats['completed'] / stats['total'] * 100) if stats['total'] > 0 else 0

            ws.cell(row=row, column=1, value=project_name)
            ws.cell(row=row, column=2, value=stats['total'])
            ws.cell(row=row, column=3, value=stats['completed'])
            ws.cell(row=row, column=4, value=stats['minutes'])
            ws.cell(row=row, column=5, value=f"{stats['avg']:.1f}")
            ws.cell(row=row, column=6, value=f"{completion_rate:.1f}%")
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width